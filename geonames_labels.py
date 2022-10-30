from selenium.webdriver import Firefox
from webdriver_manager.firefox import GeckoDriverManager
from selenium.webdriver.firefox.service import Service as FirefoxService
from selenium.webdriver.common.by import By
from time import sleep
import openpyxl
import tkinter as tk
from tkinter import filedialog

root = tk.Tk()
root.withdraw()
file = filedialog.askopenfilename() # Selezione del foglio di calcolo

fire_driver = GeckoDriverManager().install() # Installazione driver
driver = Firefox(service=FirefoxService(fire_driver))
driver.maximize_window()

wb = openpyxl.load_workbook(file) # Apertura del foglio di calcolo per lettura e scrittura
sh = wb.active
m_row = sh.max_row

def get_url(cell):
    url = sh.cell(row=i, column=1).value.replace("<", "").replace(">", "") # Impostazione dell'URL della risorsa GeoNames come variabile
    return url

def get_label(i): # Per ogni riga del foglio di calcolo viene ricercato l'URL dell'elemento GeoNames e vengono estratti Etichetta e Riferimento geografico
    url = get_url(i)
    driver.get(url)
    sleep(1)
    
    name = driver.find_element(By.XPATH,
    '/html/body/div[2]/h4').get_attribute("innerHTML").split(" - ")
    
    name_label = name[0]
    place_label = driver.find_elements(By.XPATH,
    '/html/body/div[2]/div/div[1]/div/div[4]/small/a')[-1].get_attribute("innerHTML")   
  
    if place_label==name_label: # Se il riferimento geografico è uguale all'elemento GeoNames viene riportato solo quest'ultimo
        label = name_label
    else: # In caso contrario viene creata la variabile composta secondo la struttura Etichetta <Riferimento geografico>
        label = name_label + " " + "<" + place_label + ">"    

    labelCell = sh.cell(row=i, column=3) # La variabile viene salvata all'interno della terza colonna del foglio di calcolo
    labelCell.value = label
    wb.save(file)

for i in range(1, m_row + 1): # Per ogni riga del foglio di calcolo viene ripetuta l'intera procedura
    try:
        get_label(i)
    except:
        pass
    









